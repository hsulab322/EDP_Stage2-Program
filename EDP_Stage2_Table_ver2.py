import pandas as pd
import numpy as np
import re
import os
import openpyxl as opx
from openpyxl.styles import Font, Color, Fill, PatternFill, Border, Side
from openpyxl.formula.translate import Translator


def turnintolist(df):
    df1 = pd.DataFrame(index=range(0, 9), columns=df.columns)
    for col in df.columns:
        lst = eval(df[col][0])
        for i in range(len(lst)):
            df1.loc[i, col] = lst[i]
    return df1.fillna(0).astype('int', errors='ignore')

# Prepare the data


def DataPreparation(filepath, ColumnType):
    def usecols(x): return True if ColumnType in x else False
    df_col = pd.read_csv(filepath, usecols=usecols)
    df_col = turnintolist(df_col[[x for x in df_col.columns]])
    return df_col


def AllExpData(filepath):
    Choice_Round = DataPreparation(filepath, 'Choice_Round')
    Outcome_Round = DataPreparation(filepath, 'Outcome_Round')
    x_Round = DataPreparation(filepath, 'x_Round')
    x_star_Round = DataPreparation(filepath, 'x_star_Round')

    return [Choice_Round, Outcome_Round, x_Round, x_star_Round]


class EDP_table:
    def __init__(self, filepath):
        self.filepath = filepath
        self.parentfolder = os.path.dirname(filepath)
        df = pd.read_csv(self.filepath)
        self.alpha = df['alpha'][0]
        self.beta = df['beta'][0]
        self.x8pos = df['max_gain'][0]
        self.x8neg = df['max_loss'][0]
        self.Subject = re.search(
            '(c?[\d+_\-]+)_EDP_stage2_[0-9\-]+\.csv', filepath).group(1)
        [self.Choice_Round, self.Outcome_Round, self.x_Round,
            self.x_star_Round] = AllExpData(self.filepath)

    def CreateWorkbook(self):
        # Create a workbook
        workbook = opx.Workbook()
        return workbook

    def Workbook_Merge(self):
        workbook = self.CreateWorkbook()

        # get the first worksheet
        sheet = workbook.worksheets[0]

        # All cells need to be merged

        first_merge = ['A1:B2', 'C1:D2', 'E1:E2', 'F1:F2', 'G1:G2', 'H1:H2', 'I1:I2', 'J1:J2',
                       'K1:K2', 'L1:L2', 'M1:N2', 'O1:O2', 'P1:Q2', 'R1:R2', 'S1:S2', 'T1:X2']
        first_headers = ['Subject', self.Subject, 'alpha', self.alpha, 'beta', self.beta, 'x8+', self.x8pos, 'x8-', self.x8neg, 'Performance', '',
                         'Win Rate', '', 'Note', 'Green box as quit; Black box as bankrupt']

        second_merge = [chr(i+65) + '3:' + chr(i+66) +
                        '3' for i in range(24) if i % 2 == 0]
        second_headers = ['Round', 'Endowment', 'Outcome']+[
            '1 (0.9)', '2 (0.7)', '3 (0.5)', '4 (0.3)', '5 (0.1)', '6 (0.0)', '7 (-0.1)', '8 (-0.3)', '9 (-0.5)']

        round_merge = ['A' + str(i) + ':B' + str(i+1)
                       for i in range(4, 28) if i % 2 == 0]
        round_header = list(range(1, 13))

        initial_merge = ['C' + str(i) + ':D' + str(i+1)
                         for i in range(4, 28) if i % 2 == 0]
        initial_header = list(self.x_Round.iloc[0, :].astype(int) * 2)

        outcome_merge = ['E' + str(i) + ':F' + str(i) for i in range(4, 28)]
        outcome_header = ['Gain', 'Loss'] * 12

        GLR_merge = list()
        for i in np.linspace(4, 26, 12, dtype=int):
            GLR_merge += [chr(j+72) + str(i) + ':' + chr(j+72) + str(i+1)
                          for j in range(18) if j % 2 == 0]

        all_merge = first_merge + second_merge + round_merge + \
            outcome_merge + initial_merge + GLR_merge
        all_headers = first_headers + second_headers + \
            round_header + outcome_header + initial_header

        # merge all cells
        for cells in all_merge:
            sheet.merge_cells(cells)

        # alignment
        for i in range(len(all_headers)):
            pos = all_merge[i].find(':')
            cell = sheet[all_merge[i][:pos]]
            cell.value = all_headers[i]
            cell.alignment = opx.styles.Alignment(
                horizontal='center', vertical='center')

        # borders of cells # Change all font
        cellRange = sheet['A1':'X27']
        Allfill = Side(border_style="thin", color="000000")
        for row in cellRange:
            for cell in row:
                cell.border = Border(
                    top=Allfill, right=Allfill, left=Allfill, bottom=Allfill)
                cell.font = Font(name='Times New Roman', size=12)

        # Change the font
        cellRange = first_merge + second_merge + round_merge
        for i in range(len(cellRange)):
            pos = cellRange[i].find(':')
            if cellRange[i][:pos] in ['C1', 'F1', 'H1', 'J1', 'L1', 'O1', 'R1']:
                continue
            else:
                cell = sheet[cellRange[i][:pos]]
                cell.font = Font(name='Times New Roman', bold=True, size=12)

        # change gain and loss font color
        Gaincolor = Font(name='Times New Roman', color="0000FF", size=12)
        Losscolor = Font(name='Times New Roman', color='FF0000', size=12)
        outcome_merge
        for i in range(len(outcome_merge)):
            pos = outcome_merge[i].find(':')
            cell = sheet[outcome_merge[i][:pos]]
            if cell.value == 'Gain':
                cell.font = Gaincolor
            else:
                cell.font = Losscolor

        return workbook

    def EDP_Stage2_Table(self, dir):

        workbook = self.Workbook_Merge()
        # get the first worksheet
        sheet = workbook.worksheets[0]

        # prepare some stylings

        Wincolor = Font(name='Times New Roman', color="0000FF", size=12)
        Losecolor = Font(name='Times New Roman', color='FF0000', size=12)
        Winfill = PatternFill(fgColor="CCECFF", fill_type="solid")
        Losefill = PatternFill(fgColor='FFCCFF', fill_type="solid")
        QuitBorder = Side(border_style="thick", color="009900")
        BrokeBorder = Side(border_style="thick", color="000000")
        NormalBorder = Side(border_style="thin", color="000000")

        # plug in the data and do stylings

        GainLoss = list()
        for i in range(4, 28):
            GainLoss += [chr(j+71) + str(i) for j in range(18) if j % 2 == 0]

        a = np.array(GainLoss)
        a.shape = (12, 2, 9)
        allpos = [(rnd, row, col) for rnd in range(12)
                  for col in range(9) for row in range(2)]

        Quit_count = 0
        QuitW_count = 0
        for pos in allpos:
            rnd, row, col = pos
            c = a[rnd, row, col]
            cell = sheet[c]

            if row == 0:
                val = int(self.x_Round.iloc[col, rnd])
                cell.value = int(
                    self.x_Round.iloc[col, rnd]) if val != 0 else ""
                cell.font = Wincolor
                if self.Outcome_Round.iloc[col, rnd] == 'Win':
                    cell.fill = Winfill

                if self.Choice_Round.iloc[col, rnd] == 'Quit':
                    cell.border = Border(
                        top=QuitBorder, left=QuitBorder, right=NormalBorder, bottom=NormalBorder)
                    adjcell = sheet.cell(cell.row, cell.column+1)
                    adjcell.border = Border(
                        top=QuitBorder, right=QuitBorder, left=NormalBorder, bottom=NormalBorder)
                    Quit_count += 1
                    if self.Outcome_Round.iloc[col-1, rnd] == 'Win':
                        QuitW_count += 1

                if self.Choice_Round.iloc[col, rnd] == 'Bet' and col != 8:
                    if self.Choice_Round.iloc[col+1, rnd] == 0:
                        cell.border = Border(
                            top=BrokeBorder, left=BrokeBorder, right=NormalBorder, bottom=NormalBorder)
                        adjcell = sheet.cell(cell.row, cell.column+1)
                        adjcell.border = Border(
                            top=BrokeBorder, right=BrokeBorder, left=NormalBorder, bottom=NormalBorder)

            else:
                val = int(self.x_star_Round.iloc[col, rnd])
                cell.value = int(
                    self.x_star_Round.iloc[col, rnd]) if val != 0 else ''
                cell.font = Losecolor
                if self.Outcome_Round.iloc[col, rnd] == 'Lose':
                    cell.fill = Losefill
                if self.Choice_Round.iloc[col, rnd] == 'Quit':
                    cell.border = Border(
                        bottom=QuitBorder, left=QuitBorder, right=NormalBorder, top=NormalBorder)
                    adjcell = sheet.cell(cell.row, cell.column+1)
                    adjcell.border = Border(
                        bottom=QuitBorder, right=QuitBorder, left=NormalBorder, top=NormalBorder)
                if self.Choice_Round.iloc[col, rnd] == 'Bet' and col != 8:
                    if self.Choice_Round.iloc[col+1, rnd] == 0:
                        cell.border = Border(
                            bottom=BrokeBorder, left=BrokeBorder, right=NormalBorder, top=NormalBorder)
                        adjcell = sheet.cell(cell.row, cell.column+1)
                        adjcell.border = Border(
                            bottom=BrokeBorder, right=BrokeBorder, left=NormalBorder, top=NormalBorder)

        # sheet['H4'] = "=G4/-G5"
        GLR_merge = list()
        for i in np.linspace(4, 26, 12, dtype=int):
            GLR_merge += [chr(j+72) + str(i) + ':' + chr(j+72) + str(i+1)
                          for j in range(18) if j % 2 == 0]

        for cells in GLR_merge:
            pos = cells.find(':')
            c = sheet[cells[:pos]]

            GainCell = sheet.cell(c.row, c.column-1)
            LossCell = sheet.cell(c.row+1, c.column-1)
            if GainCell.value != '':
                sheet[cells[:pos]
                      ] = f"=ROUND({GainCell.coordinate}/-{LossCell.coordinate},2)"
                sheet[cells[:pos]].alignment = opx.styles.Alignment(
                    horizontal='center', vertical='center')

        PerformanceCell = sheet['O1']
        try:
            PerformanceCell.value = pd.read_csv(
                self.filepath)['totalIncentive'][0]
            if PerformanceCell.value < 0:
                PerformanceCell.font = Font(
                    name='Times New Roman', color='FF0000', size=12)
            else:
                PerformanceCell.font = Font(
                    name='Times New Roman', color='0000FF', size=12)
        except Exception as e:
            print('error:', e)

        WinrateCell = sheet['R1']
        win = self.Outcome_Round.stack().value_counts()['Win']
        lose = self.Outcome_Round.stack().value_counts()['Lose']
        WinrateCell.value = round(win/(win+lose), 2)

        # QuitW_Cell = sheet['Q1']
        # QuitW_Cell.value = QuitW_count
        # Quit_Cell = sheet['R1']
        # Quit_Cell.value = Quit_count

        # save workbook
        TableName = self.Subject+'_Stage2_Table.xlsx'
        TableSaveName = os.path.join(dir, TableName)
        workbook.save(TableSaveName)
        print(TableSaveName, 'is done')

        return TableSaveName
